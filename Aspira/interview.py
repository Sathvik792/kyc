import pyttsx3
import speech_recognition as sr
from sentence_transformers import SentenceTransformer, util
import time
import openpyxl
import pandas as pd
from Aspira.interview_host import Host
from Aspira.interview_host import Synthesizer
import pyautogui
import os

"""
    Conducts the Interview:
    
    Reads the questions from the xlsx file and asks the quser the question and records the answer.    
    Returns:
        _type_: _description_
"""

class Interview:
    def __init__(self,questions_excel_sheet_path) -> None:
        self.player = self.set_up_player()
        self.listener = sr.Recognizer()
        self.virtual_audio_cable_index = 0
        self.model = SentenceTransformer("paraphrase-MiniLM-L6-v2")
        self.questions, self.answers = self.get_interview_question(
            questions_excel_sheet_path
        )
        # self.questions,self.answers=["what is your full name?","what is your date of birth?","where do you work?"],["sathwik reddy komandla","sept 7 2002","Walking tree technologies"]
        self.questions_asked=0
        self.started_interview = False
        self.user_answers=[]
        self.similarity_scores=[]

        #
        # Check Camera
        # Check Microphone
        # Check any other buttons

    def take_screenshot(self):
        os.makedirs("screenshots",exist_ok=True)
        screenshot=pyautogui.screenshot()
        print("took the ss")
        s_s=os.listdir("screenshots")
        screenshot.save(f"screenshots/screenshot_{len(s_s)+1}.jpg")
        # print("saved the ss")
        # screenshot.save("ss.jpg")

    def set_up_player(self):
        player = pyttsx3.init()
        player.setProperty("audioDevice", str(1))
        player.setProperty("voice", player.getProperty("voices")[1].id)
        rate = player.getProperty("rate")
        new_rate = int(rate * 0.76)
        player.setProperty("rate", new_rate)
        return player

    def talk(self, text):
        print("--------talk--------", text)
        self.player.say(text)
        self.player.runAndWait()

    def get_interview_question(self, questions_excel_sheet_path):
        # workbook = openpyxl.load_workbook("Aspira-questionaire.xlsx")
        workbook=openpyxl.load_workbook(questions_excel_sheet_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for head, val in zip(header, row):
                row_dict[head] = str(val)
            data.append(row_dict)

        questions = []
        answers = []

        for each in data:
            questions.append(each["Question"])
            answers.append(each["Answer"])
        return questions, answers

    def text_similarity(self, user_answer, csv_answer):
        user_embedding = self.model.encode([user_answer], convert_to_tensor=True)
        csv_embedding = self.model.encode([csv_answer], convert_to_tensor=True)
        similarity_score = util.pytorch_cos_sim(user_embedding, csv_embedding)
        return similarity_score.item()

    def initiate_interview(self, name):
        self.talk(
            f"Hi {name}, I am Aspira, I will be asking you some questions regarding your loan application? Are you ready for the interaction?"
        )
        while not self.started_interview:
            try:
                with sr.Microphone(
                    device_index=self.virtual_audio_cable_index
                ) as input_device:
                    print("I am ready, Listening ....")
                    voice_content = self.listener.listen(input_device, timeout=5)
                try:
                    print("main try")
                    text_command = self.listener.recognize_google(voice_content)
                    text_command = text_command.lower()
                    print(text_command, "-----start command-----")
                    if any(
                        keyword in text_command
                        for keyword in [
                            "yes",
                            "ready",
                            "iam ready",
                            "yes please start",
                            "start",
                            "lets start",
                            "please start"
                        ]
                    ):
                        self.started_interview = True

                except sr.WaitTimeoutError:
                    # self.talk(
                    #     f"Hi {name}, please say Start the Interview, When ever you are ready. Thank you"
                    # )
                    continue
                except sr.UnknownValueError:
                    print("did not understand")
                    continue
            except sr.WaitTimeoutError:
                # self.talk(f"Hi user, please say Start the Interview, When ever you are ready. Thank you")
                pass

    def ask_question(self, question)-> str:
        """
        Asks the question using Questioner and processes the responses until the question is answered.
        """
        print("   asking question        ",question)
        self.take_screenshot()
        text_command = ""
        self.talk(question)
        repeat_requested = True
        count=0
        listener = sr.Recognizer()
        while repeat_requested:
            with sr.Microphone(
                device_index=self.virtual_audio_cable_index
            ) as input_device:
                print("I am ready, Listening ....")
                try:
                    voice_content = listener.listen(input_device, timeout=5)
                    text_command = listener.recognize_google(voice_content)
                    text_command = text_command.lower()
                    if "repeat" in text_command:
                        self.talk(question)
                    else:
                        repeat_requested = False
                except sr.UnknownValueError:
                    count+=1
                    self.talk("Do you want to repeat the question ")
                    print("In unknown value error 407")
                    with sr.Microphone(
                        device_index=self.virtual_audio_cable_index
                    ) as input_device:
                        try:
                            print("I am ready, Listening ....")
                            print(time.time())
                            voice_content = listener.listen(
                                input_device, timeout=5
                            )
                            text_command = listener.recognize_google(voice_content)
                            text_command = text_command.lower()
                            print("text_command: ", text_command)
                            if (
                                text_command == "yes"
                                or "repeat the question" in text_command
                            ):
                                print("text_command: ", text_command)
                                self.talk(question)
                                with sr.Microphone() as input_device:
                                    print("I am ready, Listening ....")
                                    voice_content = listener.listen(
                                        input_device, timeout=10
                                    )
                                    text_command = listener.recognize_google(
                                        voice_content
                                    )
                                    text_command = text_command.lower()
                                    break
                            elif text_command == "no":
                                repeat_requested = False
                                break
                            else:
                                if text_command != "":
                                    break
                                else:
                                    text_command = ""
                                    repeat_requested = False
                                    break
                        except sr.WaitTimeoutError:
                            count+=1
                            print("waited long enough 443")
                            print(time.time())
                            pass
                        except sr.UnknownValueError:
                            count+=1
                            print("unknown speech 447")
                            print(time.time())
                            pass
                except sr.WaitTimeoutError:
                    count+=1
                    self.talk("Do you want me to repeat the question?")
                    print("In waittimeouterror 446")
                    with sr.Microphone(
                        device_index=self.virtual_audio_cable_index
                    ) as input_device:
                        try:
                            voice_content = listener.listen(
                                input_device, timeout=5
                            )
                            text_command = listener.recognize_google(voice_content)
                            text_command = text_command.lower()
                            print("text_command:", text_command)
                            if (
                                text_command == "yes"
                                or "repeat the question" in text_command
                            ):
                                self.talk(question)
                                with sr.Microphone() as input_device:
                                    print("I am ready, Listening .... in 195")
                                    print(time.time())
                                    voice_content = listener.listen(
                                        input_device, timeout=10
                                    )
                                    text_command = listener.recognize_google(
                                        voice_content
                                    )
                                    text_command = text_command.lower()
                                    break
                            elif text_command == "no":
                                repeat_requested = False
                                break
                            else:
                                if text_command != "":
                                    break
                                else:
                                    text_command = ""
                                    repeat_requested = False
                                    break

                        except sr.WaitTimeoutError:
                            count+=1
                            print(time.time())
                            pass

                        except sr.UnknownValueError:
                            count+=1
                            print(time.time())
                            pass
                print(
                    f"User's Response: {text_command} {len(text_command)}, count is : {count}"
                )
        return text_command

    def check_response(self, question, answer):
        relevance_response=self.host.check_relevance(
            question=question, answer=answer, main_question=question
        )
        print("releance response is -----------",relevance_response)
        if relevance_response:
            try:
                if relevance_response.get("answered"):
                    if self.follow_up:
                        self.follow_up=None
                    return answer
                else:
                    if self.follow_up:
                        return self.follow_up.start_follow_up(follow_up_question=relevance_response["follow_up_question"])
                    else:
                        self.host.chat_history=[]
                        self.follow_up = FollowUp(
                            main_question=question,
                            follow_up_question=relevance_response["follow_up_question"],
                            interview=self
                        )
                        return self.follow_up.start_follow_up()
            except Exception as e:
                print("exception answered------------------,",e)
        else:
            return answer

    def start_interview(self,name):

        #  for question, answer in zip(self.questions, self.answers):
        #     print(question, answer)
        #     self.host.chat_history = []
        #     user_response = self.ask_question(question=question)

        #     if not user_response or ("don't know" or " Next question" in user_response):
        #         answer=self.check_response(question=question,answer=user_response)
        #         # follow_up_question,follow_up_answer=ref_validator(host=evaluator,question=question,answer=user_response,virtual_audio_cable_index=virtual_audio_cable_index)
        #         # if follow_up_question and follow_up_answer:
        #         #     question_1.append(follow_up_question)
        #         #     user_responses1.append(follow_up_answer)
        #         #     print("\n\ntesting similarity score\n\n")
        #         #     similarity_score = text_similarity(user_response, answer, model)
        #         #     similarity_score = similarity_score * 100
        #         #     similarity_score = round(similarity_score)
        #         #     print(f"Similarity Score: {similarity_score}")
        #         #     user_response.append(user_response)
        #         #     user_responses1.append(user_response)
        #         #     score_response.append(similarity_score)
        #     else:
        #         # score_response.append(0)
        #         # user_response.append("")
        #         print("user didnot answer the question")
        #     self.questions_asked+=self.questions_asked
        #     if  self.questions_asked< len(self.questions):
        #         print("answer is ----------",answer)
        #         self.talk("Okay, moving on to the next question.")

        while not self.started_interview:
            print("inside to initiate")
            self.initiate_interview(name=name)

        for question, answer in zip(self.questions, self.answers):
            print(question, answer)
            questioner=Questioner(interview=self,question=question,expected_answer=answer)
            user_answered=questioner.start_question(question=question)
            print("in questioner response is    ", user_answered)
            self.user_answers.append(user_answered)
            self.talk("moving on.")

        self.create_report()

    def create_report(self):
        for question,expected_answer,user_answer in zip(self.questions,self.answers,self.user_answers):
            similarity=self.text_similarity(user_answer=user_answer,csv_answer=expected_answer)
            self.similarity_scores.append(similarity)

        FINAL_RECORD = []

        for question, answer, user_response, score in zip(
            self.questions, self.answers, self.user_answers, self.similarity_scores
        ):
            FINAL_RECORD.append(
                {
                    "Question": question,
                    "Answer": answer,
                    "User Response": user_response,
                    "%\ Match": score,
                }
            )
        df = pd.DataFrame.from_dict(FINAL_RECORD)
        df.to_excel("Interview_Summary.xlsx", index=None)

class Questioner:
    def __init__(self, interview: Interview, question, expected_answer) -> None:
        self.interview = interview
        self.question = question
        self.expected_answer = expected_answer
        self.host=Host(main_question=question)
        self.follow_ups = []

    def start_question(self, question):
        user_response = self.interview.ask_question(question=question)
        print("checking relevenace of the user response  ---   ",user_response)
        if not user_response or ("don't know" or "next question" in user_response):
            self.follow_ups.append({"Question": question, "Answer": user_response})
            user_response_eval_details = self.host.check_relevance(question=question, answer=user_response)
            try:

                if user_response_eval_details:
                   
                    print("-------------     user eval details are provided    ----------",user_response_eval_details)
                    if not user_response_eval_details["answered"] or not user_response_eval_details["follow_up_question"]:
                        self.interview.ask_question(question=user_response_eval_details["follow_up_question"])
                    else:
                        print("looks like user answrred the question",user_response_eval_details)
            except KeyError as e:
                print("key error",e)
                self.follow_ups.pop()
                return self.start_question(question=question)
            
        else:
            print("user didnot answer the question")
            return self.start_question(question=question)

        print("sysnthesizing output -------------------------------")
        print("------------",self.follow_ups)

        if self.follow_ups:
            print("#"*20,"Creating sysnthesizer")
            print(self.follow_ups)
            synthesizer=Synthesizer(main_question=question,follow_ups=self.follow_ups)
            answer=synthesizer.get_answer()
            print("sysnthesized--------------",answer)
            return answer
        else:
            return user_response


# inter=Interview()
# inter.start_interview()
